###
import numpy as np
import os
import matplotlib as plt
import matplotlib.pyplot as plt
from matplotlib import rc
from jinja2 import Template
import pathlib
from pathlib import Path
from datetime import datetime
import openpyxl


plt.rc('lines', lw=1.5)
plt.rcParams.update({'legend.fontsize':12})
plt.rcParams.update({'font.size':12})
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
plt.rcParams['xtick.major.pad'] = 10
plt.rcParams['ytick.major.pad'] = 10

cm = 1/2.54


def generate_capacitor(capacitor, vendor, f_esr = 500, SHOW_PLOT=True):


    print("\n###############################################################")
    print("############ EXTRACTING VALUES FOR: %s ############" %(capacitor))
    print("###############################################################")


    ### change working directory to where script is located:
    abspath = os.path.abspath(__file__)
    dname = os.path.dirname(abspath)
    os.chdir(dname)
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")

    capacitor = str(capacitor)

    dc_data_name = capacitor + r"_dc_bias.csv"
    ac_data_name = capacitor + r"_ac.csv"
    esr_data_name = capacitor + r"_esr.csv"

    ### check if file paths exist:
    path_parent = pathlib.Path("")
    path_dc_data = os.path.join(path_parent,r"input",vendor,capacitor,dc_data_name)
    path_ac_data = os.path.join(path_parent,r"input",vendor,capacitor,ac_data_name)
    path_esr_data = os.path.join(path_parent,r"input",vendor,capacitor,esr_data_name)

    if(Path(path_dc_data).is_file() and Path(path_ac_data).is_file() and Path(path_esr_data).is_file()):


        ### open excel database of components:
        path_to_comp = os.path.join(r"component_database",r"component_database.xlsx")

        database = openpyxl.load_workbook(path_to_comp)

        ws = database[vendor]
        ## make this into a function later: 

        for row in ws.rows:
            row_name = str(row[0].value)
            if row_name == str(capacitor):
                #capacitor value is 1st column: (scary to do it this way)
                cap_nom = row[1].value
                #voltage rating is 2nd column:
                voltage_rating = row[2].value
            
        ### make directory for output (throw error if folder exists):
        try:
            os.mkdir("output")
        except OSError as error:
            print(error)

        try:
            os.mkdir(os.path.join(path_parent,r"output",capacitor))
        except OSError as error:
            print(error)

        ## Define all the output paths to where verilog-a files are saved:
        output_path_veriloga = os.path.join(path_parent,r"output",capacitor,r"veriloga.va")
        output_path_veriloga_rc = os.path.join(path_parent,r"output",capacitor,r"veriloga_rc.va")
        output_path_veriloga_rc_500khz = os.path.join(path_parent,r"output",capacitor,r"veriloga_rc_"+str(f_esr)+r"khz.va")
        output_path_veriloga_rlc = os.path.join(path_parent,r"output",capacitor,r"veriloga_rlc.va")
        output_path_veriloga_rlc_500khz = os.path.join(path_parent,r"output",capacitor,r"veriloga_rlc_"+str(f_esr)+r"khz.va")

        ### load all data:


        dc_data = np.genfromtxt(path_dc_data, delimiter=',')
        ac_data = np.genfromtxt(path_ac_data, delimiter=',')
        esr_data = np.genfromtxt(path_esr_data, delimiter=',')

        ### handling irregular headers for csv files - not pretty but effective.
        for i in range(len(dc_data[:,0])):
            is_header = np.isnan(dc_data[i,0]).any()
            if(not is_header):
                dc_data=dc_data[i:,:]
                break
        
        for i in range(len(ac_data[:,0])):
            is_header = np.isnan(dc_data[i,0]).any()
            if(not is_header):
                ac_data=ac_data[i:,:]
                break

        for i in range(len(esr_data[:,0])):
            is_header = np.isnan(esr_data[i,0]).any()
            if(not is_header):
                esr_data=esr_data[i:,:]
                break
        

        #voltage is column 0 of dc_data:
        voltage = dc_data[:,0]

        # dc_bias degradation is column 1 and in percentage in dc_data:
        dc_deg = dc_data[:,1]/100

        #frequency for impedance meas:
        freq1 = ac_data[:,0]

        #impedance meas:
        imp_ac = ac_data[:,1]

        #frequency for resistance meas:
        freq2 = esr_data[:,0]

        #resistance meas:
        res_ac = esr_data[:,1]



        ########### Lets have fun ##############

        ##### extract ESR and inductance

        # extract esr @ f_esr:
        # find nearest index of f_esr:
        idx = (np.abs(freq2 - f_esr*1e3)).argmin()
        idx_min = res_ac.argmin()
        #take esr from resistance measurement @f_esr:
        esr_val = res_ac[idx]
        esr_val_min = np.min(res_ac)
        

        print("\nESR is: %2.4f mOhm @%2.4fkHz" %(esr_val*1e3, f_esr))
        print("\nESR_min is: %2.4f mOhm" %(esr_val_min*1e3))

        # calculate approximate L at high frequencies
        ESL = np.sqrt((imp_ac[-20]**2-esr_val**2))/(2*np.pi * freq1[-20])

        print("\nESL is: %2.3f nH \n" %(ESL*1e9))


        #### write ESR and ESL values back to excel database:
        for row in ws.rows:
            row_name = str(row[0].value)
            if row_name == capacitor:
                #capacitor value is 1st column: (scary to do it this way)
                row[3].value = esr_val_min
                #voltage rating is 2nd column:
                row[4].value = ESL

        database.save(path_to_comp)

        ##### Polynomial fit for non-linear dc behaviour:

        ### calculate 15th order polynomial fit:
        coeff = np.polyfit(voltage,dc_deg,15)
        polyfit=0

        for i in range(len(coeff)):
            polyfit = polyfit + coeff[i] * voltage**(len(coeff)-1-i)

        ##### load template

        ## first for veriloga:
        #template path:
        path_template = os.path.join(path_parent,r"templates",r"veriloga_template.txt")

        mod_template = Template(pathlib.Path(path_template).read_text())

        ### render file from template:
        mod_text = mod_template.render(date_time = dt_string, capacitor_name=capacitor, cap_nom = cap_nom, coeff=coeff, voltage_rating = voltage_rating)

        # save result to a file
        with open(f'{output_path_veriloga}', 'w') as fid:
            fid.write(mod_text)

        ## second for veriloga_rc:
        #template path:
        path_template = os.path.join(path_parent,r"templates",r"veriloga_rc_template.txt")

        mod_template = Template(pathlib.Path(path_template).read_text())

        ### render file from template:
        mod_text = mod_template.render(date_time = dt_string, capacitor_name=capacitor, cap_nom = cap_nom, coeff=coeff, esr_val=esr_val_min, voltage_rating = voltage_rating)

        # save result to a file
        with open(f'{output_path_veriloga_rc}', 'w') as fid:
            fid.write(mod_text)

        mod_text = mod_template.render(date_time = dt_string, capacitor_name=capacitor, cap_nom = cap_nom, coeff=coeff, esr_val=esr_val, voltage_rating=voltage_rating)

        # save result to a file
        with open(f'{output_path_veriloga_rc_500khz}', 'w') as fid:
            fid.write(mod_text)


        ## last for veriloga_rlc:
        #template path:
        path_template = os.path.join(path_parent,r"templates",r"veriloga_rlc_template.txt")

        mod_template = Template(pathlib.Path(path_template).read_text())

        ### render file from template:
        mod_text = mod_template.render(date_time = dt_string, capacitor_name=capacitor, cap_nom = cap_nom, coeff=coeff, esr_val=esr_val_min, ind_val = ESL, voltage_rating=voltage_rating)

        # save result to a file
        with open(f'{output_path_veriloga_rlc}', 'w') as fid:
            fid.write(mod_text)

        ### render file from template:
        mod_text = mod_template.render(date_time = dt_string, capacitor_name=capacitor, cap_nom = cap_nom, coeff=coeff, esr_val=esr_val, ind_val = ESL, voltage_rating=voltage_rating)

        # save result to a file
        with open(f'{output_path_veriloga_rlc_500khz}', 'w') as fid:
            fid.write(mod_text)

        print("#################################################")
        print("########## Script finished as expected! #########")
        print("#################################################")

        ### visualize data:
        fig, ax = plt.subplots(figsize=(15*cm,15*cm))
        ax.plot(voltage,dc_deg*100, label = "Raw data")
        ax.plot(voltage,polyfit*100,linestyle='dashdot', label = "Polynomial fit")

        ax.grid(visible=True, which='major', color='#666666', linestyle='-')
        ax.minorticks_on()
        ax.grid(visible=True, which='minor', color='#999999', linestyle='-', alpha=0.2)

        ax.set_xlabel('DC Voltage [V]')
        ax.set_ylabel('Cap. deg [%]')

        ax.legend(loc = "upper right")

        plt.tight_layout()


        ### visualize data:
        fig, ax = plt.subplots(figsize=(15*cm,15*cm))
        ax.plot(freq1,imp_ac, label = "Z ($\Omega$)")
        ax.plot(freq2,res_ac, label = "ESR ($\Omega$)")

        ax.plot(freq1[-20],imp_ac[-20], marker = '*', color='black')
        ax.plot(freq2[idx],res_ac[idx], marker = '*', color='black')
        ax.plot(freq2[idx_min],res_ac[idx_min], marker = '*', color='black')

        ESR_string = "$ESR = %2.2f m\Omega$ @ %3.2f kHz" %(esr_val*1e3, freq2[idx]*1e-3)
        ESL_string = "ESL = %2.3f nH" %(ESL*1e9)
        ESR_min_string = "$ESR_{min} = %2.2f m\Omega$ @ %3.2f kHz" %(esr_val_min*1e3, freq2[idx_min]*1e-3)

        ax.annotate(ESL_string, xy=(freq1[-20],imp_ac[-20]), xycoords='data', ha="center", va="bottom", fontsize=10)
        ax.annotate(ESR_string, xy=(freq2[idx],res_ac[idx]), xycoords='data', ha="center", va="bottom", fontsize=10)
        ax.annotate(ESR_min_string, xy=(freq2[idx_min],res_ac[idx_min]), xycoords='data', ha="center", va="top", fontsize=10)


        ax.set_yscale('log')
        ax.set_xscale('log')

        ax.grid(visible=True, which='major', color='#666666', linestyle='-')
        ax.minorticks_on()
        ax.grid(visible=True, which='minor', color='#999999', linestyle='-', alpha=0.2)

        ax.set_xlabel('Frequency [Hz]')
        ax.set_ylabel('Impedance [$\mathrm{\Omega}$]')
        ax.legend(loc = "upper right")
        plt.tight_layout()



        if SHOW_PLOT:
            plt.show()

    else:
        print("\n WarningInfo:      The csv data could not be found -- ensure that you have extracted the files in the correct folder and with the correct naming! \n")



def make_all(f_esr = 500):

    #inputs:
        #f_esr (default) = 500kHz
    ### change working directory to where script is located:
    abspath = os.path.abspath(__file__)
    dname = os.path.dirname(abspath)
    os.chdir(dname)
    ### open excel database of components:
    path_to_comp = os.path.join(r"component_database",r"component_database.xlsx")

    database = openpyxl.load_workbook(path_to_comp)

    for vendor in database.sheetnames:
        ws = database[vendor]
        for col in ws['A']:
            if (col.value is not None and col.value != "name"):
                #run generate_capacitor function:
                generate_capacitor(col.value,vendor, SHOW_PLOT = False)
                #close the plots after you:
                plt.close('all')
            



if __name__ == "__main__":

    ##################################################################################
    ################################### EXAMPLE 1: ###################################
    ##################################################################################

    ### Generate verilog-A files for a GRM21BC71E106KE11 - murata capacitor.
    ### Before this script is run the following has been done by the user:
    # 1:    The user has found the component on the vendor webpage and downloaded the
    #       the csv data for: ac impedance measurement, ac_esr measurement and the
    #       capacitance DC bias voltage degradation
    #
    # 2:    The user has inserted ATLEAST the capacitor name and nominal capacitance
    #       value in the component_database.xlsx excel file. The ESR and ESL values
    #       should not be inserted, since they are extracted from this script.
    #
    # 3:    The user has ensured that the csv data is saved under the correct names
    #       and in the correct path as specified by the user guide. 
    #
    # 4:    Remember to close the component_database.xlsx file before running the
    #       the script. Otherwise the script cannot open and edit the file.

    ######## USER INPUTS #########

    ### choose manufacture vendor:
    vendor = "murata"

    # part number:
    capacitor = "GRM21BC71E106KE11"


    # Frequency for f_esr extraction:
    f_esr = 500              #in kHz

    ### show plots or not (for visual inspection of polynomial fit and impedance extraction)
    SHOW_PLOT = True

    #Run main generator function:
    generate_capacitor(capacitor, vendor, f_esr)


    ##################################################################################
    ################################### EXAMPLE 2: ###################################
    ##################################################################################

    ### Generate verilog-A files for a all available in component_database.
    #   if errors exist in database or missing csv data, the function should
    #   report a WarningInfo and skip these.

    f_esr = 500              #in kHz
    #make_all(f_esr)